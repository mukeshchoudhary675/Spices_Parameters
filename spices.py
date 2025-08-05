import streamlit as st
import pandas as pd
from openpyxl import load_workbook

st.title("🧪 Parameter Group Analysis (from Excel Formatting)")

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])
if uploaded_file:
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    data = []
    current_group = None

    for i, row in enumerate(ws.iter_rows(min_row=2)):
        cell_values = [cell.value for cell in row]
        if row[0].font and row[0].font.bold:
            current_group = row[0].value
            continue
        if cell_values[0] and cell_values[4]:
            data.append([
                current_group,
                cell_values[0],
                cell_values[1],
                cell_values[2],
                cell_values[3],
                cell_values[4],
            ])

    df = pd.DataFrame(data, columns=[
        "Group", "Parameter", "No. of Samples tested",
        "No. of Non-compliant Samples", "% Non-compliance", "Commodity"
    ])

    st.subheader("📄 Cleaned Data")
    st.dataframe(df)

    # Convert numeric columns safely
    df["No. of Samples tested"] = pd.to_numeric(df["No. of Samples tested"], errors="coerce")
    df["No. of Non-compliant Samples"] = pd.to_numeric(df["No. of Non-compliant Samples"], errors="coerce")
    
    # Summary: Total Samples & Non-Compliant by Group
    st.subheader("📊 Summary by Group")
    df_summary = df.groupby("Group", as_index=False).agg({
        "No. of Samples tested": "sum",
        "No. of Non-compliant Samples": "sum"
    })
    df_summary["% Non-compliance"] = (
        df_summary["No. of Non-compliant Samples"] / df_summary["No. of Samples tested"]
    ).round(4)
    
    st.dataframe(df_summary)


    # Download cleaned file
    st.download_button("⬇️ Download Cleaned Data",
                       df.to_csv(index=False),
                       file_name="cleaned_parameters.csv",
                       mime="text/csv")
